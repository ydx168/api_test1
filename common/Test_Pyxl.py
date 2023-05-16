import openpyxl


class TestPyxl:
    def __init__(self, pash, sheet):
        self.pash = pash
        self.workbook = openpyxl.load_workbook(self.pash)
        self.she = self.workbook[sheet]

    def __del__(self):
        self.workbook.close()

    def all(self):
        re = list(self.she.values)
        return re

    def sheet(self):
        return self.workbook.sheetnames

    def row(self):
        return self.she.max_row

    def column(self):
        return self.she.max_column

    def add(self, name):
        self.she.append(name)
        self.workbook.save(self.pash)

    def appoint(self, r_ow, c_olumn, name):
        row = r_ow
        column = c_olumn
        value = name
        self.she.cell(row=row, column=column, value=value)
        self.workbook.save(self.pash)
